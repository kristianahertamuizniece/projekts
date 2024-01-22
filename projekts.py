import openpyxl
from openpyxl import Workbook, load_workbook
import selenium
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import time

service=Service()
option=webdriver.ChromeOptions()
option.add_argument('--headless')
driver=webdriver.Chrome(service=service, options=option)

workbook=load_workbook('gramatas.xlsx')
sheet=workbook.active

GramatuSaraksts=[]
for row in range(2, sheet.max_row+1):
    gramata=sheet.cell(row=row, column=1).value
    autors=sheet.cell(row=row,column=2).value
    GramataAutors=f"{gramata} {autors}"
    GramatuSaraksts.append(GramataAutors)

url="https://www.janisroze.lv/lv/"
driver.get(url)
time.sleep(2)


gramata=[]

for entry in GramatuSaraksts:
    find=driver.find_element(By.ID, "search")
    find.clear()
    find.send_keys(entry)

    time.sleep(4)
    cena=driver.find_element(By.CSS_SELECTOR, "span.price")
    price=cena.text

    gramata.append({'gramata':entry, 'price':price})

for item in gramata:
        print(f"{item['gramata']} {item['price']}")